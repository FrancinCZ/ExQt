import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import hashlib
import json
from pathlib import Path
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


REPORT_PRIMARY_COLUMNS = [
    "filename", "cell_id", "object_id", "Z_px", "Y_px", "X_px",
    "volume_px", "volume_bio_um3", "equivalent_diameter_um",
    "mean_intensity", "max_intensity", "integrated_density",
    "A_object", "A_shell", "A_middle", "A_core",
    "Delta_A_middle_shell", "Delta_A_core_middle", "Delta_A_core_shell",
    "mode_a_core_voxels", "mode_a_z_topology_status",
    "mode_a_z_split_slice_fraction", "mode_a_object_touches_edge",
    "mode_a_object_touches_roi_edge", "radial_monotonic",
    "primary_qc_valid",
]

REPORT_EXCLUDED_COLUMNS = [
    "filename", "cell_id", "object_id", "Z_px", "Y_px", "X_px",
    "volume_px", "volume_bio_um3", "equivalent_diameter_um",
    "mean_intensity", "A_shell", "A_middle", "A_core",
    "Delta_A_middle_shell", "Delta_A_core_middle", "Delta_A_core_shell",
    "mode_a_core_voxels", "mode_a_empty_layers",
    "mode_a_layer_complete_coverage", "mode_a_object_touches_edge",
    "mode_a_object_touches_roi_edge", "mode_a_z_topology_status",
    "mode_a_z_occupied_slices", "mode_a_z_split_slices",
    "mode_a_z_split_slice_fraction", "fa_complete", "primary_qc_valid",
    "mode_a_qc_reason",
]


def _metadata_value(metadata, dotted_path, default=None):
    value = metadata
    for key in dotted_path.split("."):
        if not isinstance(value, dict) or key not in value:
            return default
        value = value[key]
    return value


def _single_column_value(df, column, default=None):
    if column not in df.columns:
        return default
    values = df[column].dropna().unique().tolist()
    return values[0] if len(values) == 1 else default


def _load_run_metadata(csv_path):
    metadata_path = csv_path.with_name(f"{csv_path.stem}_metadata.json")
    if not metadata_path.exists():
        return {}, metadata_path
    with metadata_path.open("r", encoding="utf-8") as handle:
        return json.load(handle), metadata_path


# Calibration is deliberately shown but excluded from the compatibility
# fingerprint: correctly calibrated acquisitions may use different sampling.
def _qc_policy_entries(metadata, df, min_size, max_size):
    fields = [
        ("analysis", "mode", _metadata_value(metadata, "parameters.mode", _single_column_value(df, "mode")), True,
        "Analysis dimensionality."),
        ("analysis", "expansion_factor", _metadata_value(metadata, "parameters.expansion_factor"), True,
        "Biological scale correction."),
        ("analysis", "min_voxels", _metadata_value(metadata, "parameters.min_voxels"), True,
        "Minimum raw connected-component size."),
        ("analysis", "plot_min_size", min_size, True,
        "Lower biological-size bound used for eligible and primary objects."),
        ("analysis", "plot_max_size", max_size, True,
        "Upper biological-size bound used for eligible and primary objects."),
        ("channels", "signal_channel", _metadata_value(metadata, "channels.signal_channel"), True,
        "Signal channel index."),
        ("channels", "dapi_channel", _metadata_value(metadata, "channels.dapi_channel"), True,
        "DAPI channel index."),
        ("mode_a", "enabled", _metadata_value(metadata, "mode_a.enabled", "A_shell" in df.columns), True,
        "Whether Radial FA Profiling metrics are active."),
        ("mode_a", "min_core_voxels", _metadata_value(
            metadata, "mode_a.min_core_voxels", _single_column_value(df, "mode_a_min_core_voxels")
        ), True, "Minimum voxels required for each FA layer and core."),
        ("mode_a", "require_z_topology_pass_for_primary", _metadata_value(
            metadata, "mode_a.require_z_topology_pass_for_primary"
        ), True, "Only Z-topology PASS objects may enter the primary set."),
        ("mode_a", "z_split_policy", _metadata_value(metadata, "mode_a.z_split_policy"), True,
        "Versioned Z-topology policy."),
        ("mode_a", "z_split_min_component_voxels", _metadata_value(
            metadata, "mode_a.z_split_min_component_voxels",
            _single_column_value(df, "mode_a_z_split_min_component_voxels")
        ), True, "Absolute minimum for a substantial separated component."),
        ("mode_a", "z_split_min_component_fraction", _metadata_value(
            metadata, "mode_a.z_split_min_component_fraction",
            _single_column_value(df, "mode_a_z_split_min_component_fraction")
        ), True, "Relative minimum for a substantial separated component."),
        ("mode_a", "z_split_pass_fraction", _metadata_value(
            metadata, "mode_a.z_split_pass_fraction",
            _single_column_value(df, "mode_a_z_split_pass_fraction")
        ), True, "Maximum split-slice fraction classified as PASS."),
        ("mode_a", "z_split_review_fraction", _metadata_value(
            metadata, "mode_a.z_split_review_fraction",
            _single_column_value(df, "mode_a_z_split_review_fraction")
        ), True, "Maximum split-slice fraction classified as REVIEW; higher is FAIL."),
        ("mode_a", "layer_scheme", _metadata_value(
            metadata, "mode_a.layer_scheme", _single_column_value(df, "mode_a_layer_scheme")
        ), True, "Radial shell-middle-core partition scheme."),
        ("mode_a", "sampling_order", _metadata_value(
            metadata, "mode_a.sampling_order", _single_column_value(df, "mode_a_sampling_order")
        ), True, "Axis order used by the anisotropy calculation."),
        ("calibration", "pixel_size_nm", _metadata_value(metadata, "parameters.pixel_size_nm"), False,
        "Run-specific physical calibration; may legitimately differ."),
        ("calibration", "z_step_nm", _metadata_value(metadata, "parameters.z_step_nm"), False,
        "Run-specific physical calibration; may legitimately differ."),
        ("provenance", "timestamp", _metadata_value(metadata, "timestamp"), False,
        "Run timestamp; not part of policy compatibility."),
        ("provenance", "software", _metadata_value(metadata, "software"), False,
        "Software name; not part of policy compatibility."),
    ]
    return fields


def qc_policy_fingerprint(metadata, df, min_size=None, max_size=None):
    if min_size is None:
        min_size = _metadata_value(metadata, "parameters.plot_min_size")
    if max_size is None:
        max_size = _metadata_value(metadata, "parameters.plot_max_size")
    entries = _qc_policy_entries(metadata, df, min_size, max_size)
    payload = {
        f"{section}.{parameter}": value
        for section, parameter, value, included, _ in entries
        if included
    }
    canonical = json.dumps(payload, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest(), entries


def _prepare_reporting_frames(df, min_size, max_size):
    report_df = df.copy()
    if "volume_bio_um3" in report_df.columns:
        size_col = "volume_bio_um3"
        report_df["equivalent_diameter_um"] = ((6 * report_df[size_col]) / np.pi) ** (1 / 3)
    elif "area_bio_um2" in report_df.columns:
        size_col = "area_bio_um2"
        report_df["equivalent_diameter_um"] = 2 * np.sqrt(report_df[size_col] / np.pi)
    elif "shape_metric_bio" in report_df.columns:
        size_col = "shape_metric_bio"
        is_3d = bool(report_df["is_3d"].iloc[0]) if "is_3d" in report_df.columns else True
        if is_3d:
            report_df["equivalent_diameter_um"] = ((6 * report_df[size_col]) / np.pi) ** (1 / 3)
        else:
            report_df["equivalent_diameter_um"] = 2 * np.sqrt(report_df[size_col] / np.pi)
    else:
        raise ValueError("Missing columns for biological size calculation.")

    report_df["equivalent_diameter_nm"] = report_df["equivalent_diameter_um"] * 1000.0
    size_values = pd.to_numeric(report_df[size_col], errors="coerce")
    size_eligible = size_values.ge(float(min_size)) & size_values.le(float(max_size))

    mode_a_available = {
        "A_object_valid", "A_shell_valid", "A_middle_valid", "A_core_valid",
        "A_shell", "A_middle", "A_core", "mode_a_primary_include",
    }.issubset(report_df.columns)
    if mode_a_available:
        complete_fa = pd.Series(True, index=report_df.index)
        for column in ("A_object_valid", "A_shell_valid", "A_middle_valid", "A_core_valid"):
            complete_fa &= _as_bool_series(report_df[column])
        complete_fa &= report_df[["A_shell", "A_middle", "A_core"]].apply(
            pd.to_numeric, errors="coerce"
        ).notna().all(axis=1)
        primary_requested = _as_bool_series(report_df["mode_a_primary_include"])
    else:
        complete_fa = pd.Series(True, index=report_df.index)
        primary_requested = pd.Series(True, index=report_df.index)

    primary_mask = size_eligible & complete_fa & primary_requested
    report_df["size_eligible"] = size_eligible
    report_df["fa_complete"] = complete_fa
    report_df["primary_qc_valid"] = primary_mask
    if {"A_shell", "A_middle", "A_core"}.issubset(report_df.columns):
        report_df["radial_monotonic"] = (
            pd.to_numeric(report_df["A_shell"], errors="coerce")
            <= pd.to_numeric(report_df["A_middle"], errors="coerce")
        ) & (
            pd.to_numeric(report_df["A_middle"], errors="coerce")
            <= pd.to_numeric(report_df["A_core"], errors="coerce")
        )
    else:
        report_df["radial_monotonic"] = pd.NA

    primary = report_df.loc[primary_mask, [
        column for column in REPORT_PRIMARY_COLUMNS if column in report_df.columns
    ]].copy()
    excluded = report_df.loc[size_eligible & ~primary_mask, [
        column for column in REPORT_EXCLUDED_COLUMNS if column in report_df.columns
    ]].copy()
    # This is a derived audit view, not a replacement for the source batch CSV.
    # Keep the explicit reporting flags so every classification can be traced.
    raw = report_df.copy()
    return raw, primary, excluded, report_df, size_col


def _count_formula(sheet_name, frame, column):
    column_letter = get_column_letter(frame.columns.get_loc(column) + 1)
    last_row = max(2, len(frame) + 1)
    return f"COUNTA('{sheet_name}'!${column_letter}$2:${column_letter}${last_row})"


def _column_range(sheet_name, frame, column):
    column_letter = get_column_letter(frame.columns.get_loc(column) + 1)
    last_row = max(2, len(frame) + 1)
    return f"'{sheet_name}'!${column_letter}$2:${column_letter}${last_row}"


def _style_table_sheet(ws, status_column=None):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34
    for column_cells in ws.iter_cols(1, ws.max_column):
        header = str(column_cells[0].value or "")
        width = min(max(len(header) + 2, 11), 24)
        if header in {"filename", "mode_a_qc_reason", "mode_a_empty_layers", "mode_a_layer_scheme"}:
            width = 28
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    if status_column and status_column in [cell.value for cell in ws[1]]:
        col = [cell.value for cell in ws[1]].index(status_column) + 1
        fills = {
            "pass": PatternFill("solid", fgColor="E2F0D9"),
            "review": PatternFill("solid", fgColor="FFF2CC"),
            "fail": PatternFill("solid", fgColor="FCE4D6"),
        }
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            fill = fills.get(str(cell.value).strip().lower())
            if fill:
                cell.fill = fill


def _build_per_cell_summary(report_df, size_col):
    rows = []
    for (filename, cell_id), group in report_df.groupby(["filename", "cell_id"], dropna=False):
        eligible = group.loc[group["size_eligible"]]
        primary = group.loc[group["primary_qc_valid"]]
        rows.append({
            "filename": filename,
            "cell_id": cell_id,
            "raw_segmented_components": len(group),
            "size_eligible_objects": len(eligible),
            "complete_fa_objects": int(eligible["fa_complete"].sum()),
            "primary_qc_valid_objects": len(primary),
            "qc_acceptance_rate": (len(primary) / len(eligible)) if len(eligible) else np.nan,
            "primary_median_size": primary[size_col].median() if len(primary) else np.nan,
            "primary_median_delta_a": primary["Delta_A_core_shell"].median()
            if len(primary) and "Delta_A_core_shell" in primary.columns else np.nan,
        })
    return pd.DataFrame(rows)


def _build_standard_cell_counts(raw_df, size_eligible_df):
    """Return per-cell raw and size-eligible counts for standard plots.

    Radial FA Profiling QC is deliberately not consulted here: these are descriptive
    counts for the standard report, with the size filter made explicit.
    """
    cell_keys = ["filename", "cell_id"]
    counts_all = (
        raw_df.groupby(cell_keys, dropna=False)
        .size()
        .reset_index(name="all_segmented_count")
    )
    counts_eligible = (
        size_eligible_df.groupby(cell_keys, dropna=False)
        .size()
        .reset_index(name="size_eligible_count")
    )
    counts = counts_all.merge(counts_eligible, on=cell_keys, how="left")
    counts["size_eligible_count"] = counts["size_eligible_count"].fillna(0).astype(int)
    return counts


def _write_summary_sheet(writer, raw, primary, excluded, report_df, size_col, fingerprint, min_size, max_size):
    workbook = writer.book
    ws = workbook.create_sheet("Summary", 0)
    writer.sheets["Summary"] = ws
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    ws["A1"] = "ExQt analysis summary"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="17365D")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    section_fill = PatternFill("solid", fgColor="D9EAF7")
    for range_name, title in (("A3:C3", "QC funnel"), ("E3:G3", "Primary-object statistics"),
                            ("A14:C14", "QC exclusion overview"), ("E14:G14", "Run configuration")):
        ws.merge_cells(range_name)
        cell = ws[range_name.split(":")[0]]
        cell.value = title
        cell.font = Font(bold=True, color="17365D")
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")

    raw_count = _count_formula("All_Objects_Raw", raw, "object_id")
    primary_count = _count_formula("Primary_Objects", primary, "object_id")
    excluded_count = _count_formula("QC_Excluded", excluded, "object_id")
    fa_complete_ranges = []
    for sheet_name, frame in (("Primary_Objects", primary), ("QC_Excluded", excluded)):
        if "fa_complete" in frame.columns:
            rng = _column_range(sheet_name, frame, "fa_complete")
            fa_complete_ranges.append(f'COUNTIF({rng},1)')
        elif sheet_name == "Primary_Objects":
            fa_complete_ranges.append(_count_formula(sheet_name, frame, "object_id"))
    complete_formula = "+".join(fa_complete_ranges) or "0"
    if size_col == "volume_bio_um3":
        size_name, size_unit = "biological volume", "µm³"
    elif size_col == "area_bio_um2":
        size_name, size_unit = "biological area", "µm²"
    else:
        size_name, size_unit = "biological size", "calibrated units"
    metrics = [
        ("Raw segmented components", f"={raw_count}", "All connected components before biological-size and QC filtering."),
        ("Size-eligible objects", f"={primary_count}+{excluded_count}", f"Objects within {min_size:g}–{max_size:g} {size_unit}."),
        ("Complete FA objects", f"={complete_formula}", "Size-eligible objects with valid shell, middle and core FA."),
        ("Primary QC-valid objects", f"={primary_count}", "Objects used in primary Radial FA Profiling graphs and statistics."),
        ("QC acceptance rate", "=IFERROR(B8/B6,0)", "Primary QC-valid divided by size-eligible; raw noise is not the denominator."),
    ]
    ws.append([])
    ws["A4"], ws["B4"], ws["C4"] = "Metric", "Value", "Definition"
    for row_index, (label, formula, definition) in enumerate(metrics, start=5):
        ws.cell(row=row_index, column=1, value=label)
        ws.cell(row=row_index, column=2, value=formula)
        ws.cell(row=row_index, column=3, value=definition)
    ws["B9"].number_format = "0.0%"

    primary_stats = []
    if size_col in primary.columns:
        rng = _column_range("Primary_Objects", primary, size_col)
        primary_stats.extend([
            (f"Median {size_name} ({size_unit})", f'=IFERROR(MEDIAN({rng}),"")', "Primary objects only."),
            (f"Mean {size_name} ({size_unit})", f'=IFERROR(AVERAGE({rng}),"")', "Primary objects only."),
        ])
    if "equivalent_diameter_um" in primary.columns:
        rng = _column_range("Primary_Objects", primary, "equivalent_diameter_um")
        primary_stats.append(("Median equivalent diameter (µm)", f'=IFERROR(MEDIAN({rng}),"")', "Primary objects only."))
    if "mean_intensity" in primary.columns:
        rng = _column_range("Primary_Objects", primary, "mean_intensity")
        primary_stats.append(("Mean intensity (a.u.)", f'=IFERROR(AVERAGE({rng}),"")', "Primary objects only."))
    if "Delta_A_core_shell" in primary.columns:
        rng = _column_range("Primary_Objects", primary, "Delta_A_core_shell")
        primary_stats.extend([
            ("Mean Delta A (core − shell)", f'=IFERROR(AVERAGE({rng}),"")', "Geometric FA difference."),
            ("Median Delta A (core − shell)", f'=IFERROR(MEDIAN({rng}),"")', "Geometric FA difference."),
            ("Fraction with Delta A > 0", f'=IFERROR(COUNTIF({rng},">0")/{primary_count},0)', "Direction only; not proof of LLPS."),
        ])
    if "radial_monotonic" in primary.columns:
        rng = _column_range("Primary_Objects", primary, "radial_monotonic")
        primary_stats.append(("Monotonic shell ≤ middle ≤ core", f'=IFERROR(COUNTIF({rng},1)/{primary_count},0)', "Fraction of primary objects."))
    ws["E4"], ws["F4"], ws["G4"] = "Metric", "Value", "Definition"
    for row_index, (label, formula, definition) in enumerate(primary_stats[:8], start=5):
        ws.cell(row=row_index, column=5, value=label)
        ws.cell(row=row_index, column=6, value=formula)
        ws.cell(row=row_index, column=7, value=definition)
    for row_index in (11, 12):
        ws.cell(row=row_index, column=6).number_format = "0.0%"

    def _combined_countif(column, criterion, quote=True):
        terms = []
        for sheet_name, frame in (("Primary_Objects", primary), ("QC_Excluded", excluded)):
            if column in frame.columns:
                excel_criterion = f'"{criterion}"' if quote else str(criterion)
                terms.append(f'COUNTIF({_column_range(sheet_name, frame, column)},{excel_criterion})')
        return "+".join(terms) or "0"

    qc_rows = [
        ("Z topology PASS", f'={_combined_countif("mode_a_z_topology_status", "pass")}', "Among size-eligible objects."),
        ("Z topology REVIEW", f'={_combined_countif("mode_a_z_topology_status", "review")}', "Excluded from primary analysis."),
        ("Z topology FAIL", f'={_combined_countif("mode_a_z_topology_status", "fail")}', "Excluded from primary analysis."),
        ("Touches image edge", f'={_combined_countif("mode_a_object_touches_edge", 1, quote=False)}', "Hard QC exclusion."),
        ("Touches ROI edge", f'={_combined_countif("mode_a_object_touches_roi_edge", 1, quote=False)}', "Hard QC exclusion."),
    ]
    ws["A15"], ws["B15"], ws["C15"] = "Condition", "Count", "Meaning"
    for row_index, (label, formula, definition) in enumerate(qc_rows, start=16):
        ws.cell(row=row_index, column=1, value=label)
        ws.cell(row=row_index, column=2, value=formula)
        ws.cell(row=row_index, column=3, value=definition)

    filename = str(raw["filename"].iloc[0]) if len(raw) and "filename" in raw.columns else ""
    calibration_z = _single_column_value(raw, "mode_a_sampling_z_nm")
    calibration_y = _single_column_value(raw, "mode_a_sampling_y_nm")
    config_rows = [
        ("Input file", filename, "Source image recorded in the object table."),
        ("QC fingerprint", f"{fingerprint[:16]}…", "Full fingerprint is recorded on the QC_Policy sheet."),
        ("Eligible size range", f"{min_size:g}–{max_size:g} {size_unit}", "Run-specific reporting range."),
        ("Effective sampling Z (nm)", calibration_z, "After expansion correction; may differ between acquisitions."),
        ("Effective sampling Y/X (nm)", calibration_y, "After expansion correction; may differ between acquisitions."),
    ]
    ws["E15"], ws["F15"], ws["G15"] = "Parameter", "Value", "Meaning"
    for row_index, (label, value, definition) in enumerate(config_rows, start=16):
        ws.cell(row=row_index, column=5, value=label)
        ws.cell(row=row_index, column=6, value=value)
        ws.cell(row=row_index, column=7, value=definition)

    per_cell = _build_per_cell_summary(report_df, size_col)
    start_row = 23
    per_cell_columns = (1, 2, 3, 5, 6, 7, 8, 9, 10)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
    ws.cell(start_row, 1, "Per-file / per-cell summary")
    ws.cell(start_row, 1).font = Font(bold=True, color="17365D")
    ws.cell(start_row, 1).fill = section_fill
    for col_index, name in zip(per_cell_columns, per_cell.columns):
        ws.cell(start_row + 1, col_index, name)
    for row_index, values in enumerate(per_cell.itertuples(index=False, name=None), start=start_row + 2):
        for col_index, value in zip(per_cell_columns, values):
            ws.cell(row_index, col_index, None if pd.isna(value) else value)

    for header_row in (4, 15, start_row + 1):
        for cell in ws[header_row]:
            if cell.column <= 10 and cell.value is not None:
                cell.fill = PatternFill("solid", fgColor="5B9BD5")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A4"
    widths = {"A": 30, "B": 18, "C": 54, "D": 3, "E": 31, "F": 22, "G": 48,
            "H": 20, "I": 22, "J": 22}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in range(5, 13):
        ws.cell(row=row, column=6).number_format = "0.0000"
    for row in (11, 12):
        ws.cell(row=row, column=6).number_format = "0.0%"
    ws["B9"].number_format = "0.0%"


def _write_qc_policy_sheet(writer, entries, fingerprint, metadata_path):
    policy_df = pd.DataFrame(entries, columns=["section", "parameter", "value", "in_fingerprint", "notes"])
    policy_df["in_fingerprint"] = policy_df["in_fingerprint"].map({True: "yes", False: "no"})
    policy_df.to_excel(writer, sheet_name="QC_Policy", index=False, startrow=4)
    ws = writer.sheets["QC_Policy"]
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    ws["A1"] = "QC policy and compatibility fingerprint"
    ws["A1"].font = Font(size=15, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="17365D")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"], ws["B2"] = "QC fingerprint", fingerprint
    ws["A3"], ws["B3"] = "Source metadata", str(metadata_path)
    ws["D2"] = "Compatibility rule"
    ws["E2"] = "Fingerprints must match before pooling runs under one QC policy. Calibration rows may differ."
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[5].height = 34
    ws.auto_filter.ref = f"A5:E{ws.max_row}"
    ws.freeze_panes = "A6"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 68
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 62


def generate_excel_stats(
    csv_filename,
    min_size=None,
    max_size=None,
    *,
    generate_excel=True,
    generate_primary_csv=True,
    generate_excluded_csv=True,
    generate_raw_audit_csv=False,
):
    csv_path = Path(csv_filename)
    if not csv_path.exists():
        print(f"Error: File '{csv_filename}' not found!")
        return None

    folder_name = csv_path.resolve().parent.name
    output_excel = csv_path.parent / f"{folder_name}_Detailed_Stats.xlsx"
    output_csv = csv_path.parent / f"{folder_name}_All_Condensates_With_Diameters.csv"
    output_primary_csv = csv_path.parent / f"{folder_name}_Primary_Condensates.csv"
    output_excluded_csv = csv_path.parent / f"{folder_name}_QC_Excluded.csv"

    df = pd.read_csv(csv_path)
    metadata, metadata_path = _load_run_metadata(csv_path)
    if min_size is None:
        min_size = _metadata_value(metadata, "parameters.plot_min_size", 0.0001)
    if max_size is None:
        max_size = _metadata_value(metadata, "parameters.plot_max_size", 2.0)

    raw, primary, excluded, report_df, size_col = _prepare_reporting_frames(
        df, float(min_size), float(max_size)
    )
    fingerprint, policy_entries = qc_policy_fingerprint(
        metadata, df, float(min_size), float(max_size)
    )

    # The original batch CSV remains untouched. Every additional file is an
    # explicitly selected derived view; the redundant full-width audit copy is
    # off by default.
    if generate_raw_audit_csv:
        raw.to_csv(output_csv, index=False)
        print(f"Raw audit CSV with diameters generated: {output_csv}")
    if generate_primary_csv:
        primary.to_csv(output_primary_csv, index=False)
        print(f"Primary QC-valid CSV generated: {output_primary_csv}")
    if generate_excluded_csv:
        excluded.to_csv(output_excluded_csv, index=False)
        print(f"QC-excluded CSV generated: {output_excluded_csv}")

    if generate_excel:
        with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
            primary.to_excel(writer, sheet_name="Primary_Objects", index=False)
            excluded.to_excel(writer, sheet_name="QC_Excluded", index=False)
            raw.to_excel(writer, sheet_name="All_Objects_Raw", index=False)
            _write_qc_policy_sheet(writer, policy_entries, fingerprint, metadata_path)
            _write_summary_sheet(
                writer, raw, primary, excluded, report_df, size_col,
                fingerprint, float(min_size), float(max_size),
            )
            _style_table_sheet(writer.sheets["Primary_Objects"], "mode_a_z_topology_status")
            _style_table_sheet(writer.sheets["QC_Excluded"], "mode_a_z_topology_status")
            _style_table_sheet(writer.sheets["All_Objects_Raw"], "mode_a_z_topology_status")

            if "Delta_A_core_shell" in primary.columns and len(primary):
                delta_col = get_column_letter(primary.columns.get_loc("Delta_A_core_shell") + 1)
                delta_range = f"{delta_col}2:{delta_col}{len(primary) + 1}"
                writer.sheets["Primary_Objects"].conditional_formatting.add(
                    delta_range,
                    CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FCE4D6")),
                )
        print(f"Readable Excel report generated: {output_excel}")

    return {
        "excel": output_excel if generate_excel else None,
        "all_objects_csv": output_csv if generate_raw_audit_csv else None,
        "primary_csv": output_primary_csv if generate_primary_csv else None,
        "excluded_csv": output_excluded_csv if generate_excluded_csv else None,
        "qc_fingerprint": fingerprint,
        "counts": {
            "raw": len(raw),
            "size_eligible": len(primary) + len(excluded),
            "complete_fa": int(report_df.loc[report_df["size_eligible"], "fa_complete"].sum()),
            "primary": len(primary),
        },
    }


class QCPolicyMismatchError(ValueError):
    """Raised when selected runs cannot be pooled under one QC policy."""


def _fingerprint_payload(entries):
    return {
        f"{section}.{parameter}": value
        for section, parameter, value, included, _ in entries
        if included
    }


def find_merge_source_csvs(folder):
    root = Path(folder)
    if not root.exists() or not root.is_dir():
        raise ValueError(f"Merge folder does not exist: {root}")
    return sorted(
        path for path in root.rglob("*_Output_Batch_*.csv")
        if not any(token in path.name for token in (
            "_Primary_Condensates", "_QC_Excluded",
            "_All_Condensates_With_Diameters", "Merged_",
        ))
    )


def _run_id(root, csv_path):
    relative = csv_path.relative_to(root)
    #The containing result folder is normally the clearest stable run label. Fall back to the CSV stem only when files live directly in the root.
    if relative.parent != Path("."):
        return str(relative.parent).replace("\\", "/")
    return relative.stem


def _merge_policy_comparison(runs):
    keys = sorted({key for run in runs for key in run["policy"].keys()})
    rows = []
    for key in keys:
        values = [run["policy"].get(key) for run in runs]
        canonical = [json.dumps(value, sort_keys=True, default=str) for value in values]
        rows.append({
            "parameter": key,
            "compatible": "yes" if len(set(canonical)) == 1 else "no",
            **{run["run_id"]: value for run, value in zip(runs, values)},
        })
    return pd.DataFrame(rows)


def _style_merge_summary(ws):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "ExQt merged-run summary"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="17365D")
    ws["A1"].alignment = Alignment(horizontal="center")
    for row in (3, 12):
        for cell in ws[row]:
            if cell.value is not None:
                cell.fill = PatternFill("solid", fgColor="5B9BD5")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A4"
    widths = {"A": 34, "B": 24, "C": 20, "D": 30, "E": 22, "F": 48}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _write_merge_summary(writer, run_stats, primary, fingerprint, root):
    ws = writer.book.create_sheet("Merge_Summary", 0)
    writer.sheets["Merge_Summary"] = ws
    ws.append([])
    ws.append([])
    ws.append(["Metric", "Value", "Unit / status", "Interpretation", "Scope", "Notes"])
    positive_fraction = (
        pd.to_numeric(primary.get("Delta_A_core_shell"), errors="coerce").gt(0).mean()
        if len(primary) and "Delta_A_core_shell" in primary.columns else np.nan
    )
    monotonic_fraction = (
        _as_bool_series(primary["radial_monotonic"]).mean()
        if len(primary) and "radial_monotonic" in primary.columns else np.nan
    )
    rows = [
        ("Included runs", len(run_stats), "runs", "Independent image/run units", "run level", "Use runs—not condensates—as n for inference."),
        ("Merged primary objects", len(primary), "objects", "Descriptive pooled object count", "object level", "Not an independent biological n."),
        ("QC fingerprint", f"{fingerprint[:16]}…", "MATCH", "All included QC-relevant settings are identical", "policy", "Full fingerprint is available in Per_Run_Stats."),
        ("Median run acceptance", run_stats["qc_acceptance_rate"].median(), "fraction", "Median primary / size-eligible", "run level", ""),
        ("Median pooled Delta A", primary["Delta_A_core_shell"].median() if "Delta_A_core_shell" in primary else np.nan, "FA difference", "Core − shell", "object level", "Descriptive only."),
        ("Pooled fraction Delta A > 0", positive_fraction, "fraction", "Direction of radial FA change", "object level", "Not proof of LLPS."),
        ("Pooled monotonic fraction", monotonic_fraction, "fraction", "Shell ≤ middle ≤ core", "object level", "Not proof of LLPS."),
    ]
    for row in rows:
        ws.append(list(row))
    ws.append([])
    ws.append(["Run ID", "Primary", "Size eligible", "Acceptance", "Median Delta A", "Calibration / provenance"])
    for record in run_stats.to_dict("records"):
        calibration = (
            f"XY={record.get('pixel_size_nm', np.nan):g} nm; "
            f"Z={record.get('z_step_nm', np.nan):g} nm"
        )
        ws.append([
            record["run_id"], record["primary_objects"], record["size_eligible_objects"],
            record["qc_acceptance_rate"], record["median_delta_a"], calibration,
        ])
    ws["A2"] = "Source folder"
    ws["B2"] = str(root)
    ws.merge_cells("B2:F2")
    for row in range(4, 11):
        if ws.cell(row, 1).value in {"Median run acceptance", "Pooled fraction Delta A > 0", "Pooled monotonic fraction"}:
            ws.cell(row, 2).number_format = "0.0%"
    for row in range(12, ws.max_row + 1):
        ws.cell(row, 4).number_format = "0.0%"
        ws.cell(row, 5).number_format = "0.0000"
    _style_merge_summary(ws)


def _compact_run_labels(run_ids):
    """Return readable, unique labels without exposing full absolute paths."""
    labels = []
    used = set()
    for index, run_id in enumerate(run_ids, start=1):
        parts = Path(str(run_id).replace("/", "\\")).parts
        candidate = "/".join(parts[-2:]) if len(parts) >= 2 else str(run_id)
        if candidate in used:
            candidate = f"R{index}: {candidate}"
        used.add(candidate)
        labels.append(candidate)
    return labels


def _generate_merge_plot(output_path, run_stats, primary):
    plot_path = Path(output_path).with_suffix(".png")
    run_ids = run_stats["run_id"].astype(str).tolist()
    labels = _compact_run_labels(run_ids)
    color_values = [
        "#4E79A7", "#F28E2B", "#59A14F", "#E15759", "#B07AA1",
        "#76B7B2", "#EDC948", "#FF9DA7", "#9C755F", "#BAB0AC",
    ]
    colors = [color_values[index % len(color_values)] for index in range(len(run_ids))]

    def font(size, bold=False):
        filename = "arialbd.ttf" if bold else "arial.ttf"
        try:
            return ImageFont.truetype(str(Path("C:/Windows/Fonts") / filename), size)
        except OSError:
            return ImageFont.load_default()

    image = Image.new("RGB", (1800, 1280), "white")
    draw = ImageDraw.Draw(image)
    title_font, panel_font, body_font, small_font = font(32, True), font(22, True), font(17), font(14)
    draw.text((900, 30), "ExQt merged-run overview", fill="#17365D", font=title_font, anchor="ma")
    draw.text(
        (900, 72), "Run-level summaries; pooled condensates are descriptive only",
        fill="#555555", font=body_font, anchor="ma",
    )

    panels = [(80, 135, 850, 580), (950, 135, 1720, 580),
            (80, 665, 850, 1110), (950, 665, 1720, 1110)]

    def frame(panel, heading, y_min, y_max, y_label, x_labels):
        left, top, right, bottom = panel
        draw.text((left, top - 35), heading, fill="#222222", font=panel_font)
        plot = (left + 90, top + 20, right - 25, bottom - 75)
        x0, y0, x1, y1 = plot
        draw.rectangle(plot, outline="#777777", width=2)
        for fraction in np.linspace(0, 1, 6):
            y = int(y1 - fraction * (y1 - y0))
            value = y_min + fraction * (y_max - y_min)
            draw.line((x0, y, x1, y), fill="#E1E1E1", width=1)
            draw.text((x0 - 10, y), f"{value:.2f}", fill="#444444", font=small_font, anchor="rm")
        draw.text((left + 12, (top + bottom) / 2), y_label, fill="#333333", font=small_font, anchor="mm")
        positions = np.linspace(x0 + 30, x1 - 30, max(len(x_labels), 1))
        for x, label in zip(positions, x_labels):
            draw.text((int(x), y1 + 15), label, fill="#333333", font=small_font, anchor="ma")
        return plot, positions

    def y_position(value, bounds, plot):
        _, y0, _, y1 = plot
        low, high = bounds
        if high <= low or pd.isna(value):
            return None
        clipped = min(max(float(value), low), high)
        return int(y1 - (clipped - low) / (high - low) * (y1 - y0))

    layer_columns = ["A_shell", "A_middle", "A_core"]
    plot, x_positions = frame(
        panels[0], "A) Per-run median radial FA", 0, 1,
        "Median FA", ["Shell", "Middle", "Core"],
    )
    if len(primary) and set(layer_columns).issubset(primary.columns):
        medians = primary.groupby("run_id", sort=False)[layer_columns].median()
        for index, run_id in enumerate(run_ids):
            if run_id not in medians.index:
                continue
            points = []
            for x, value in zip(x_positions, medians.loc[run_id]):
                y = y_position(value, (0, 1), plot)
                if y is not None:
                    points.append((int(x), y))
            if len(points) >= 2:
                draw.line(points, fill=colors[index], width=3)
            for point in points:
                draw.ellipse((point[0] - 6, point[1] - 6, point[0] + 6, point[1] + 6), fill=colors[index])

    delta = pd.to_numeric(run_stats.get("median_delta_a"), errors="coerce")
    finite_delta = delta[np.isfinite(delta)]
    delta_low = min(-0.05, float(finite_delta.min()) - 0.03) if len(finite_delta) else -0.1
    delta_high = max(0.05, float(finite_delta.max()) + 0.03) if len(finite_delta) else 0.1
    short_labels = [f"R{index + 1}" for index in range(len(run_ids))]
    plot, x_positions = frame(
        panels[1], "B) Per-run median Delta A", delta_low, delta_high,
        "Median core - shell", short_labels,
    )
    zero_y = y_position(0, (delta_low, delta_high), plot)
    if zero_y is not None:
        draw.line((plot[0], zero_y, plot[2], zero_y), fill="#555555", width=2)
    for index, (x, value) in enumerate(zip(x_positions, delta)):
        y = y_position(value, (delta_low, delta_high), plot)
        if y is not None:
            draw.ellipse((int(x) - 8, y - 8, int(x) + 8, y + 8), fill=colors[index])

    plot, x_positions = frame(
        panels[2], "C) QC acceptance by run", 0, 1,
        "Primary / size-eligible", short_labels,
    )
    acceptance = pd.to_numeric(run_stats["qc_acceptance_rate"], errors="coerce")
    bar_width = max(18, min(70, int((plot[2] - plot[0]) / max(len(run_ids), 1) * 0.55)))
    for index, (x, value) in enumerate(zip(x_positions, acceptance)):
        y = y_position(value, (0, 1), plot)
        if y is None:
            continue
        draw.rectangle((int(x) - bar_width // 2, y, int(x) + bar_width // 2, plot[3]), fill=colors[index])
        record = run_stats.iloc[index]
        draw.text(
            (int(x), max(y - 8, plot[1] + 8)),
            f"{int(record['primary_objects'])}/{int(record['size_eligible_objects'])}",
            fill="#222222", font=small_font, anchor="ms",
        )

    object_delta = pd.to_numeric(primary.get("Delta_A_core_shell"), errors="coerce")
    finite_object = object_delta[np.isfinite(object_delta)] if len(primary) else pd.Series(dtype=float)
    object_low = min(-0.05, float(finite_object.min()) - 0.03) if len(finite_object) else -0.1
    object_high = max(0.05, float(finite_object.max()) + 0.03) if len(finite_object) else 0.1
    plot, x_positions = frame(
        panels[3], "D) QC-valid objects by run (descriptive)", object_low, object_high,
        "Object core - shell", short_labels,
    )
    zero_y = y_position(0, (object_low, object_high), plot)
    if zero_y is not None:
        draw.line((plot[0], zero_y, plot[2], zero_y), fill="#555555", width=2)
    if len(primary) and "Delta_A_core_shell" in primary.columns:
        for index, run_id in enumerate(run_ids):
            values = pd.to_numeric(
                primary.loc[primary["run_id"].astype(str) == run_id, "Delta_A_core_shell"],
                errors="coerce",
            ).dropna().tolist()
            for point_index, value in enumerate(values):
                jitter = ((point_index * 37) % 21 - 10) * 1.6
                x = int(x_positions[index] + jitter)
                y = y_position(value, (object_low, object_high), plot)
                draw.ellipse((x - 4, y - 4, x + 4, y + 4), fill=colors[index])

    legend_y = 1150
    for index, label in enumerate(labels):
        column = index % 3
        row = index // 3
        x, y = 100 + column * 570, legend_y + row * 30
        draw.rectangle((x, y + 3, x + 18, y + 21), fill=colors[index])
        draw.text((x + 28, y), f"R{index + 1}: {label}", fill="#333333", font=small_font)

    image.save(plot_path, format="PNG", optimize=True)
    return plot_path


def merge_statistics_folder(folder, output_path=None, include_raw=False):

    root = Path(folder).resolve()
    csv_paths = find_merge_source_csvs(root)
    if not csv_paths:
        raise ValueError("No original *_Output_Batch_*.csv files were found in the selected folder.")

    runs = []
    for csv_path in csv_paths:
        metadata, metadata_path = _load_run_metadata(csv_path)
        if not metadata_path.exists():
            raise ValueError(f"Missing metadata for {csv_path}: expected {metadata_path.name}")
        df = pd.read_csv(csv_path)
        min_size = _metadata_value(metadata, "parameters.plot_min_size")
        max_size = _metadata_value(metadata, "parameters.plot_max_size")
        if min_size is None or max_size is None:
            raise ValueError(f"Missing analyzed biological size range in {metadata_path}")
        raw, primary, excluded, report_df, size_col = _prepare_reporting_frames(
            df, float(min_size), float(max_size)
        )
        fingerprint, entries = qc_policy_fingerprint(metadata, df, float(min_size), float(max_size))
        run_id = _run_id(root, csv_path)
        for frame in (raw, primary, excluded, report_df):
            frame.insert(0, "source_csv", str(csv_path))
            frame.insert(0, "run_id", run_id)
        runs.append({
            "run_id": run_id,
            "csv_path": csv_path,
            "metadata_path": metadata_path,
            "metadata": metadata,
            "fingerprint": fingerprint,
            "entries": entries,
            "policy": _fingerprint_payload(entries),
            "raw": raw,
            "primary": primary,
            "excluded": excluded,
            "report": report_df,
            "size_col": size_col,
        })

    comparison = _merge_policy_comparison(runs)
    incompatible = comparison.loc[comparison["compatible"] == "no"]
    if len(incompatible):
        details = []
        for _, row in incompatible.iterrows():
            values = ", ".join(
                f"{run['run_id']}={row[run['run_id']]}" for run in runs
            )
            details.append(f"{row['parameter']}: {values}")
        raise QCPolicyMismatchError(
            "Selected runs do not share one QC policy:\n" + "\n".join(details)
        )

    merged_primary = pd.concat([run["primary"] for run in runs], ignore_index=True, sort=False)
    merged_excluded = pd.concat([run["excluded"] for run in runs], ignore_index=True, sort=False)
    merged_raw = pd.concat([run["raw"] for run in runs], ignore_index=True, sort=False)
    run_rows = []
    for run in runs:
        report = run["report"]
        primary = run["primary"]
        eligible = int(report["size_eligible"].sum())
        run_rows.append({
            "run_id": run["run_id"],
            "source_csv": str(run["csv_path"]),
            "source_metadata": str(run["metadata_path"]),
            "raw_components": len(report),
            "size_eligible_objects": eligible,
            "complete_fa_objects": int(report.loc[report["size_eligible"], "fa_complete"].sum()),
            "primary_objects": len(primary),
            "qc_acceptance_rate": len(primary) / eligible if eligible else np.nan,
            "median_size": primary[run["size_col"]].median() if len(primary) else np.nan,
            "median_delta_a": primary["Delta_A_core_shell"].median()
            if len(primary) and "Delta_A_core_shell" in primary.columns else np.nan,
            "fraction_delta_a_positive": pd.to_numeric(
                primary.get("Delta_A_core_shell"), errors="coerce"
            ).gt(0).mean() if len(primary) and "Delta_A_core_shell" in primary.columns else np.nan,
            "pixel_size_nm": _metadata_value(run["metadata"], "parameters.pixel_size_nm"),
            "z_step_nm": _metadata_value(run["metadata"], "parameters.z_step_nm"),
            "expansion_factor": _metadata_value(run["metadata"], "parameters.expansion_factor"),
            "qc_fingerprint": run["fingerprint"],
        })
    run_stats = pd.DataFrame(run_rows)

    output_path = Path(output_path) if output_path else root / "Merged_Stats.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        run_stats.to_excel(writer, sheet_name="Per_Run_Stats", index=False)
        merged_primary.to_excel(writer, sheet_name="Merged_Primary", index=False)
        merged_excluded.to_excel(writer, sheet_name="Merged_QC_Excluded", index=False)
        comparison.to_excel(writer, sheet_name="QC_Policy_Comparison", index=False)
        if include_raw:
            merged_raw.to_excel(writer, sheet_name="Merged_Raw_Audit", index=False)
        _write_merge_summary(writer, run_stats, merged_primary, runs[0]["fingerprint"], root)
        _style_table_sheet(writer.sheets["Per_Run_Stats"])
        _style_table_sheet(writer.sheets["Merged_Primary"], "mode_a_z_topology_status")
        _style_table_sheet(writer.sheets["Merged_QC_Excluded"], "mode_a_z_topology_status")
        _style_table_sheet(writer.sheets["QC_Policy_Comparison"])
        writer.sheets["Per_Run_Stats"].column_dimensions["A"].width = 26
        writer.sheets["Per_Run_Stats"].column_dimensions["B"].width = 60
        writer.sheets["Per_Run_Stats"].column_dimensions["C"].width = 60
        writer.sheets["Per_Run_Stats"].column_dimensions["O"].width = 68
        policy_ws = writer.sheets["QC_Policy_Comparison"]
        policy_ws.column_dimensions["A"].width = 46
        policy_ws.column_dimensions["B"].width = 14
        for column_index in range(3, policy_ws.max_column + 1):
            policy_ws.column_dimensions[get_column_letter(column_index)].width = 28
        if include_raw:
            _style_table_sheet(writer.sheets["Merged_Raw_Audit"], "mode_a_z_topology_status")

    plot_path = _generate_merge_plot(output_path, run_stats, merged_primary)

    return {
        "excel": output_path,
        "plot": plot_path,
        "run_count": len(runs),
        "primary_count": len(merged_primary),
        "excluded_count": len(merged_excluded),
        "raw_count": len(merged_raw),
        "qc_fingerprint": runs[0]["fingerprint"],
        "source_csvs": csv_paths,
    }


def generate_plots(csv_filename, min_size=0.0001, max_size=2.0):
    sns.set_theme(style="ticks", palette="muted")
    plt.rcParams.update({"font.sans-serif": "DejaVu Sans", "font.family": "sans-serif"})

    csv_path = Path(csv_filename)
    if not csv_path.exists():
        print(f"Error: File '{csv_filename}' not found!")
        return

    folder_name = csv_path.resolve().parent.name
    df = pd.read_csv(csv_path)


    if "volume_bio_um3" in df.columns:
        df["size_plot"] = df["volume_bio_um3"]
        xlabel_text = "Biological Volume (µm³)"
        title_a = "A) Condensate Volume Distribution"
        title_c = "C) Volume vs. Total Intensity"
        unit_label = "µm³"
    elif "area_bio_um2" in df.columns:
        df["size_plot"] = df["area_bio_um2"]
        xlabel_text = "Biological Area (µm²)"
        title_a = "A) Condensate Area Distribution"
        title_c = "C) Area vs. Total Intensity"
        unit_label = "µm²"
    elif "shape_metric_bio" in df.columns:
        df["size_plot"] = df["shape_metric_bio"]
        is_3d_mode = df["is_3d"].iloc[0] if "is_3d" in df.columns else True
        unit_label = "µm³" if is_3d_mode else "µm²"
        dim_name = "Volume" if is_3d_mode else "Area"
        xlabel_text = f"Biological {dim_name} ({unit_label})"
        title_a = f"A) Condensate {dim_name} Distribution"
        title_c = f"C) {dim_name} vs. Total Intensity"
    else:
        print("Error: Missing columns for size calculation!")
        return

    df_filtered = df[(df["size_plot"] >= min_size) & (df["size_plot"] <= max_size)].copy()

    # Standard plots are descriptive and intentionally independent of Radial FA Profiling
    # QC.  Show both the segmented-component count and the biologically sized
    # subset so the size filter cannot silently look like a low cell count.
    counts_per_cell = _build_standard_cell_counts(df, df_filtered)

    if counts_per_cell["filename"].nunique(dropna=False) == 1:
        counts_per_cell["display_label"] = counts_per_cell["cell_id"].map(
            lambda value: f"Cell {value}"
        )
    else:
        counts_per_cell["display_label"] = counts_per_cell.apply(
            lambda row: f"{Path(str(row['filename'])).stem}\nCell {row['cell_id']}",
            axis=1,
        )

    fig, axes = plt.subplots(2, 3, figsize=(22, 12))

    def show_empty_size_panel(axis, title):
        axis.set_title(title, fontsize=13, fontweight="bold")
        axis.text(
            0.5,
            0.5,
            f"No objects within\n{min_size:g}–{max_size:g} {unit_label}",
            ha="center",
            va="center",
            transform=axis.transAxes,
            fontsize=12,
        )
        axis.set_xticks([])
        axis.set_yticks([])

    if df_filtered.empty:
        show_empty_size_panel(axes[0, 0], title_a)
    else:
        sns.histplot(df_filtered["size_plot"], kde=True, ax=axes[0, 0], color="#2b5c8f", bins=30 if len(df_filtered) > 30 else 10)
        median_size = df_filtered["size_plot"].median()
        axes[0, 0].axvline(median_size, color="#d9534f", linestyle="--", linewidth=2, label=f"Median: {median_size:.4f} {unit_label}")
        axes[0, 0].set_title(title_a, fontsize=13, fontweight="bold")
        axes[0, 0].set_xlabel(xlabel_text)
        axes[0, 0].set_ylabel("Count")
        axes[0, 0].legend()

    cell_axis = axes[0, 1]
    positions = np.arange(len(counts_per_cell))
    bar_width = 0.38
    bars_all = cell_axis.bar(
        positions - bar_width / 2,
        counts_per_cell["all_segmented_count"],
        width=bar_width,
        color="#9aa9bd",
        label="All segmented components",
    )
    bars_eligible = cell_axis.bar(
        positions + bar_width / 2,
        counts_per_cell["size_eligible_count"],
        width=bar_width,
        color="#8e44ad",
        label=f"Within {min_size:g}–{max_size:g} {unit_label}",
    )
    for bars in (bars_all, bars_eligible):
        for bar in bars:
            height = int(bar.get_height())
            cell_axis.annotate(
                str(height),
                (bar.get_x() + bar.get_width() / 2, height),
                xytext=(0, 3),
                textcoords="offset points",
                ha="center",
                va="bottom",
                fontsize=8,
            )
    cell_axis.set_xticks(positions)
    cell_axis.set_xticklabels(counts_per_cell["display_label"])
    if len(counts_per_cell) > 8:
        cell_axis.tick_params(axis="x", labelrotation=45, labelsize=8)
    cell_axis.set_title("B) Objects per Cell Label", fontsize=13, fontweight="bold")
    cell_axis.set_xlabel("ROI / cell label")
    cell_axis.set_ylabel("Object count")
    cell_axis.legend(fontsize=9)

    if df_filtered.empty:
        show_empty_size_panel(axes[0, 2], title_c)
        show_empty_size_panel(axes[1, 0], "D) Signal Concentration (Mean Intensity)")
        show_empty_size_panel(axes[1, 1], "E) Size vs. Mean Intensity")
        show_empty_size_panel(axes[1, 2], "F) 2D Density (Cluster Topography)")
    else:
        sns.scatterplot(data=df_filtered, x="size_plot", y="integrated_density", ax=axes[0, 2], color="#2e7d32", s=40, alpha=0.5)
        sns.regplot(data=df_filtered, x="size_plot", y="integrated_density", ax=axes[0, 2], scatter=False, color="#1b5e20")
        axes[0, 2].set_title(title_c, fontsize=13, fontweight="bold")
        axes[0, 2].set_xlabel(xlabel_text)
        axes[0, 2].set_ylabel("Integrated Density (a.u.)")

        sns.violinplot(y=df_filtered["mean_intensity"], ax=axes[1, 0], color="#f57c00", inner="quartile", alpha=0.7)
        sns.stripplot(y=df_filtered["mean_intensity"], ax=axes[1, 0], color="black", alpha=0.2, jitter=0.15, size=3)
        axes[1, 0].set_title("D) Signal Concentration (Mean Intensity)", fontsize=13, fontweight="bold")
        axes[1, 0].set_ylabel("Mean Intensity (a.u.)")

        sns.scatterplot(data=df_filtered, x="size_plot", y="mean_intensity", ax=axes[1, 1], color="#c0392b", s=40, alpha=0.5)
        sns.regplot(data=df_filtered, x="size_plot", y="mean_intensity", ax=axes[1, 1], scatter=False, color="#922b21")
        axes[1, 1].set_title("E) Size vs. Mean Intensity", fontsize=13, fontweight="bold")
        axes[1, 1].set_xlabel(xlabel_text)
        axes[1, 1].set_ylabel("Mean Intensity (a.u.)")

        sns.kdeplot(data=df_filtered, x="size_plot", y="mean_intensity", ax=axes[1, 2], fill=True, cmap="YlOrBr", thresh=0.05, alpha=0.8)
        axes[1, 2].set_title("F) 2D Density (Cluster Topography)", fontsize=13, fontweight="bold")
        axes[1, 2].set_xlabel(xlabel_text)
        axes[1, 2].set_ylabel("Mean Intensity (a.u.)")


    plt.tight_layout()
    output_plot = csv_path.parent / f"{folder_name}_Analysis_Plots.png"
    plt.savefig(output_plot, dpi=300)
    plt.close(fig)

    print(f"Graphs were successfuly generated to: '{output_plot}'")


#Interpret CSV booleans safely after pandas reloads them as text or bool.
def _as_bool_series(series):
    return series.astype(str).str.strip().str.lower().isin({"true", "1", "yes", "on"})


#Create a separate radial FA and QC report without changing standard ExQt plots.
def generate_rezim_a_plots(csv_filename, min_size=None, max_size=None):
    sns.set_theme(style="whitegrid", palette="colorblind")
    plt.rcParams.update({"font.sans-serif": "DejaVu Sans", "font.family": "sans-serif"})

    csv_path = Path(csv_filename)
    if not csv_path.exists():
        print(f"Error: File '{csv_filename}' not found!")
        return None

    df = pd.read_csv(csv_path)
    required_columns = {
        "A_shell", "A_middle", "A_core", "Delta_A_core_shell", "mode_a_primary_include",
        "A_object_valid", "A_shell_valid", "A_middle_valid", "A_core_valid",
        "mode_a_qc_reason",
    }
    missing_columns = sorted(required_columns.difference(df.columns))
    if missing_columns:
        print("Radial FA Profiling plots skipped: missing columns: " + ", ".join(missing_columns))
        return None

    #Primary plots intentionally include only complete and QC-approved FA rows
    valid_flags = np.ones(len(df), dtype=bool)
    for column in ("A_object_valid", "A_shell_valid", "A_middle_valid", "A_core_valid"):
        valid_flags &= _as_bool_series(df[column]).to_numpy()

    numeric_columns = ["A_shell", "A_middle", "A_core", "Delta_A_core_shell"]
    for column in numeric_columns:
        df[column] = pd.to_numeric(df[column], errors="coerce")

    complete_fa = valid_flags & df[numeric_columns].notna().all(axis=1).to_numpy()
    primary_include = _as_bool_series(df["mode_a_primary_include"]).to_numpy()

    size_eligible = np.ones(len(df), dtype=bool)
    if min_size is not None or max_size is not None:
        size_column = next(
            (name for name in ("volume_bio_um3", "area_bio_um2", "shape_metric_bio") if name in df.columns),
            None,
        )
        if size_column is None:
            raise ValueError("Radial FA Profiling size filtering requested, but no biological size column is present.")
        size_values = pd.to_numeric(df[size_column], errors="coerce")
        if min_size is not None:
            size_eligible &= size_values.ge(float(min_size)).fillna(False).to_numpy()
        if max_size is not None:
            size_eligible &= size_values.le(float(max_size)).fillna(False).to_numpy()

    primary = df.loc[size_eligible & complete_fa & primary_include].copy()
    primary["Delta_A_middle_shell"] = primary["A_middle"] - primary["A_shell"]
    primary["Delta_A_core_middle"] = primary["A_core"] - primary["A_middle"]

    fig, axes = plt.subplots(2, 2, figsize=(15, 11))
    fig.suptitle(
        "Radial FA Profiling: Shell-Middle-Core Fractional Anisotropy and QC Overview\n"
        "Radial-layer panels contain only size-eligible, complete, QC-approved objects.",
        fontsize=15,
        fontweight="bold",
    )

    def no_primary_data(ax, title):
        ax.set_title(title, fontsize=12, fontweight="bold")
        ax.text(
            0.5, 0.5,
            "No primary QC-valid objects\nfor this batch.",
            ha="center", va="center", transform=ax.transAxes, fontsize=12,
        )
        ax.set_xticks([])
        ax.set_yticks([])

    paired_ax = axes[0, 0]
    paired_ax.set_title("A) Per-object Shell → Middle → Core FA", fontsize=12, fontweight="bold")
    if primary.empty:
        no_primary_data(paired_ax, "A) Per-object Shell → Middle → Core FA")
    else:
        rng = np.random.default_rng(0)
        for _, row in primary.iterrows():
            paired_ax.plot(
                [0, 1, 2],
                [row["A_shell"], row["A_middle"], row["A_core"]],
                color="#6c757d",
                alpha=0.45,
                linewidth=1,
            )
        paired_ax.scatter(rng.normal(0, 0.025, len(primary)), primary["A_shell"], color="#0f766e", label="Shell", zorder=3)
        paired_ax.scatter(rng.normal(1, 0.025, len(primary)), primary["A_middle"], color="#d97706", label="Middle", zorder=3)
        paired_ax.scatter(rng.normal(2, 0.025, len(primary)), primary["A_core"], color="#c2410c", label="Core", zorder=3)
        paired_ax.set_xticks([0, 1, 2], ["Shell", "Middle", "Core"])
        paired_ax.set_ylabel("Fractional Anisotropy (FA)")
        paired_ax.set_ylim(0, 1)
        paired_ax.legend(frameon=True)

    scatter_ax = axes[0, 1]
    scatter_ax.set_title("B) Layer-to-layer FA changes", fontsize=12, fontweight="bold")
    if primary.empty:
        no_primary_data(scatter_ax, "B) Layer-to-layer FA changes")
    else:
        for _, row in primary.iterrows():
            scatter_ax.plot(
                [0, 1],
                [row["Delta_A_middle_shell"], row["Delta_A_core_middle"]],
                color="#94a3b8",
                alpha=0.4,
                linewidth=1,
            )
        rng = np.random.default_rng(1)
        scatter_ax.scatter(
            rng.normal(0, 0.025, len(primary)),
            primary["Delta_A_middle_shell"],
            s=42,
            color="#d97706",
            alpha=0.82,
        )
        scatter_ax.scatter(
            rng.normal(1, 0.025, len(primary)),
            primary["Delta_A_core_middle"],
            s=42,
            color="#c2410c",
            alpha=0.82,
        )
        scatter_ax.axhline(0, linestyle="--", color="#6b7280", linewidth=1)
        scatter_ax.set_xticks([0, 1], ["Middle − Shell", "Core − Middle"])
        scatter_ax.set_ylabel("FA difference")

    delta_ax = axes[1, 0]
    delta_ax.set_title("C) Delta A = Core FA − Shell FA", fontsize=12, fontweight="bold")
    if primary.empty:
        no_primary_data(delta_ax, "C) Delta A = Core FA − Shell FA")
    else:
        sns.histplot(primary["Delta_A_core_shell"], bins=min(20, max(5, len(primary))), kde=len(primary) >= 5,
                    color="#7c3aed", ax=delta_ax)
        delta_ax.axvline(0, color="#374151", linestyle="--", linewidth=1)
        delta_ax.set_xlabel("Delta A (geometric difference)")
        delta_ax.set_ylabel("Object count")

    qc_ax = axes[1, 1]
    qc_ax.set_title("D) Radial FA Profiling QC Funnel", fontsize=12, fontweight="bold")
    qc_counts = pd.Series(
        [len(df), int(size_eligible.sum()), int((size_eligible & complete_fa).sum()), len(primary)],
        index=[
            "All segmented\nobjects",
            "Size-eligible\nset",
            "Complete FA\nset",
            "Primary\nQC-valid",
        ],
    )
    bars = qc_ax.bar(
        qc_counts.index,
        qc_counts.values,
        color=["#94a3b8", "#60a5fa", "#f59e0b", "#16a34a"],
    )
    qc_ax.set_ylabel("Object count")
    for bar, value in zip(bars, qc_counts.values):
        qc_ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height(), str(value), ha="center", va="bottom")

    excluded_reason_values = (
        df.loc[size_eligible & ~(complete_fa & primary_include), "mode_a_qc_reason"]
        .astype("string")
        .fillna("")
    )
    excluded_reasons = excluded_reason_values.str.split(";").explode()
    excluded_reasons = excluded_reasons[excluded_reasons != ""]
    if not excluded_reasons.empty:
        top_reasons = excluded_reasons.value_counts().head(3)
        reason_text = "Top QC reasons:\n" + "\n".join(f"{name}: {count}" for name, count in top_reasons.items())
        qc_ax.text(1.04, 0.96, reason_text, transform=qc_ax.transAxes, va="top", fontsize=9)

    fig.tight_layout(rect=(0, 0, 1, 0.93))
    output_plot = csv_path.parent / f"{csv_path.stem}_Radial_FA_Profiling_Plots.png"
    fig.savefig(output_plot, dpi=300, bbox_inches="tight")
    plt.close(fig)
    print(f"Radial FA Profiling plots were generated to: '{output_plot}'")
    return output_plot
